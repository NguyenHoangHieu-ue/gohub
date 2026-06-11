"""
sync_data_files.py — Unified file-based import pipeline.

Cơ chế:
  - Quét tất cả file trong Data/ (ref/ và ncc/ subfolder)
  - Tính SHA-256 hash của từng file
  - So sánh với hash lần import trước (lưu trong bảng data_file_registry)
  - Chỉ import file nào thay đổi (hoặc chưa import lần nào)
  - Sau khi import thành công: cập nhật hash + timestamp

Usage:
  set SUPABASE_URL=https://wfuigmfnfcijkvylrwzz.supabase.co
  set SUPABASE_SERVICE_KEY=<key>
  python sync/sync_data_files.py [--force]

Arguments:
  --force   Bỏ qua kiểm tra hash, re-import tất cả file
"""

import os, sys, hashlib, argparse
from pathlib import Path
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr); sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("ERROR: pip install supabase", file=sys.stderr); sys.exit(1)

BASE_DIR  = Path(__file__).parent.parent  # repo root (sync/ is 1 level below)
DATA_DIR  = BASE_DIR / "Data"
CHUNK     = 500
NOW       = lambda: datetime.now(timezone.utc).isoformat()


# ─── Utilities ────────────────────────────────────────────────────────────────

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def upsert_batch(sb, table: str, rows: list, pk: str):
    now = NOW()
    for i in range(0, len(rows), CHUNK):
        batch = [{**r, "synced_at": now} if "synced_at" in (list(rows[0].keys()) if rows else []) else r
                 for r in rows[i:i + CHUNK]]
        sb.table(table).upsert(batch, on_conflict=pk).execute()


def load_xlsx_rows(path: Path, min_row: int = 2):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else f"col{i}"
               for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if not any(v for v in row):
            continue
        rows.append({headers[i]: (str(v).strip() if v is not None else None)
                     for i, v in enumerate(row)})
    return rows


# ─── Import functions (one per data type) ────────────────────────────────────

def import_countries(sb, path: Path) -> int:
    """ref_countries: code, name (nameVn)"""
    raw = load_xlsx_rows(path)
    rows = []
    for r in raw:
        code = (r.get("code") or r.get("Code") or "").strip()
        name = (r.get("name") or r.get("Name") or "").strip()
        name_vn = (r.get("nameVn") or r.get("name_vn") or r.get("NameVn") or "").strip() or None
        if not code:
            continue
        rows.append({"code": code, "name": name, "name_vn": name_vn})
    sb.table("ref_countries").delete().neq("code", "PLACEHOLDER_NEVER_EXISTS").execute()
    upsert_batch(sb, "ref_countries", rows, "code")
    return len(rows)


def import_support_countries(sb, path: Path) -> int:
    """ref_support_countries: code, name, supportCountry, supportCountryVn, countryCodes"""
    raw = load_xlsx_rows(path)
    rows = []
    for r in raw:
        code = (r.get("code") or r.get("Code") or "").strip()
        if not code:
            continue
        rows.append({
            "code":               code,
            "name":               (r.get("name") or code).strip(),
            "support_country":    r.get("supportCountry") or r.get("support_country"),
            "support_country_vn": r.get("supportCountryVn") or r.get("support_country_vn"),
            "country_codes":      r.get("countryCodes") or r.get("country_codes"),
        })
    sb.table("ref_support_countries").delete().neq("code", "PLACEHOLDER_NEVER_EXISTS").execute()
    upsert_batch(sb, "ref_support_countries", rows, "code")
    return len(rows)


def import_vendors(sb, path: Path) -> int:
    """ref_vendors: Vendor Code, Name, Description"""
    raw = load_xlsx_rows(path)
    rows = []
    for r in raw:
        code = (r.get("Vendor Code") or r.get("vendor_code") or "").strip()
        name = (r.get("Name") or r.get("name") or "").strip()
        if not code:
            continue
        rows.append({
            "vendor_code": code,
            "name":        name,
            "description": r.get("Description") or r.get("description"),
        })
    upsert_batch(sb, "ref_vendors", rows, "vendor_code")
    return len(rows)


def import_categories(sb, path: Path) -> int:
    """ref_categories: category_code, name_en, name_vn, iso_code"""
    raw = load_xlsx_rows(path)
    rows = []
    for r in raw:
        code = (r.get("category_code") or r.get("Category Code") or "").strip()
        if not code:
            continue
        rows.append({
            "category_code": code,
            "name_en":       r.get("name_en") or r.get("Name EN"),
            "name_vn":       r.get("name_vn") or r.get("Name VN"),
            "iso_code":      r.get("iso_code") or r.get("ISO Code"),
            "region_type":   r.get("region_type") or "country",
            "notes":         r.get("notes"),
        })
    upsert_batch(sb, "ref_categories", rows, "category_code")
    return len(rows)


def import_fx_rates(sb, path: Path) -> int:
    """
    Đọc sheet 'FX' hoặc 'Tỷ giá' từ file tỷ giá nội bộ.
    Lấy tháng gần nhất có dữ liệu.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # Tìm sheet phù hợp
    sheet_name = next((s for s in wb.sheetnames if "FX" in s.upper() or "TY GIA" in s.upper()
                       or "TỶ GIÁ" in s), None)
    if not sheet_name:
        print(f"  [fx_rates] Không tìm thấy sheet FX trong {path.name}")
        return 0

    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(min_row=1, values_only=True))

    # Row 3 (index 2) là header: Pháp Nhân, Chỉ Tiêu, T01/2026, T02/2026...
    # Tìm cột tháng mới nhất có giá trị
    header_row = next((i for i, r in enumerate(rows_data) if r and r[0] and
                       ("phap nhan" in str(r[0]).lower() or "pháp nhân" in str(r[0]).lower()
                        or "gohub" in str(r[0]).lower())), None)
    if header_row is None:
        print(f"  [fx_rates] Không parse được header tháng")
        return 0

    headers = rows_data[header_row]
    # Tìm cột tháng gần nhất (cột bắt đầu từ index 2, format T06/2026)
    month_cols = [(i, str(h)) for i, h in enumerate(headers) if h and "T0" in str(h).upper() or
                  (h and str(h).startswith("T") and "/" in str(h))]

    # KEY_MAP: nhận dạng row theo tên chỉ tiêu → app_settings key
    KEY_MAP = {
        "vnd/usd":   ("fx.usd_vnd",  None),         # giá trị gốc VND per 1 USD
        "vnd/cny":   ("fx.vnd_cny",  None),
        "vnd/gbp":   ("fx.vnd_gbp",  None),
        "hkd/usd":   ("fx.hkd_usd",  lambda v: round(1/float(v), 6) if float(v) else None),  # 1/7.798
        "jpy/usd":   ("fx.usd_jpy",  None),
        "thb/usd":   ("fx.usd_thb",  None),
        "cny/usd":   ("fx.usd_cny",  None),
        "eur/usd":   ("fx.usd_eur",  None),
        "gbp/usd":   ("fx.usd_gbp",  None),
        "sgd/usd":   ("fx.usd_sgd",  None),
        "twd/usd":   ("fx.twd_usd",  lambda v: round(1/float(v), 6) if float(v) else None),  # 1/31.452
    }

    settings = []
    for row in rows_data[header_row + 1:]:
        if not row or not row[1]:
            continue
        label = str(row[1]).strip().lower().replace(" ", "").replace("tỷgiá", "").replace("tygia", "")
        # Extract pair like "vnd/usd"
        matched_key = next((k for k in KEY_MAP if k.replace("/", "") in label.replace("/", "")), None)
        if not matched_key:
            continue

        # Find latest non-null month value
        val = None
        month_label = None
        for col_i, col_h in reversed(month_cols):
            if col_i < len(row) and row[col_i] is not None:
                try:
                    val = float(row[col_i])
                    month_label = col_h
                    break
                except (ValueError, TypeError):
                    continue

        if val is None:
            continue

        db_key, transform = KEY_MAP[matched_key]
        final_val = transform(val) if transform else val

        settings.append({
            "key":      db_key,
            "value":    str(final_val),
            "label":    f"{matched_key.upper()} ({month_label})",
            "category": "fx_rate",
        })

    if not settings:
        return 0

    for s in settings:
        sb.table("app_settings").upsert(
            {**s, "updated_at": NOW()},
            on_conflict="key"
        ).execute()

    return len(settings)


def import_worldmove_ncc(sb, path: Path) -> int:
    """
    Import WORLDMOVE products từ xlsx.
    Cần column mapping theo format file WM.
    Đây là thin wrapper — chi tiết logic giữ trong import_worldmove_v2.py.
    """
    print("  [ncc/worldmove] Dùng import_worldmove_v2.py để import WM data")
    print("  Chạy: python database/import/import_worldmove_v2.py")
    return 0


def import_3hk_ncc(sb, path: Path) -> int:
    """Import 3HK zones từ xlsx."""
    print("  [ncc/3hk] Dùng import_3hk.py để import 3HK data")
    print("  Chạy: python database/import/import_3hk.py")
    return 0


# ─── File registry ────────────────────────────────────────────────────────────

# Map: (subfolder, filename_pattern) → import_function
FILE_MAP = {
    "countries":         import_countries,
    "support_countries": import_support_countries,
    "vendors":           import_vendors,
    "categories":        import_categories,
    "fx_rates":          import_fx_rates,
    "fx":                import_fx_rates,        # alias
    "worldmove":         import_worldmove_ncc,
    "3hk":               import_3hk_ncc,
    "3hk_zones":         import_3hk_ncc,
}


def resolve_file_key(path: Path) -> str | None:
    """Derive file_key từ tên file, bỏ qua date suffix và extension."""
    stem = path.stem.lower()
    # Bỏ date suffix dạng -2026-06-05
    import re
    stem = re.sub(r"-\d{4}-\d{2}-\d{2}$", "", stem)
    stem = stem.replace("-", "_").replace(" ", "_")
    # Bỏ tiếng Việt phức tạp
    stem = stem.replace("tỉ_giá_nội_bộ_theo_tháng", "fx_rates")
    stem = stem.replace("ty_gia_noi_bo_theo_thang", "fx_rates")
    return stem


def get_importer(file_key: str):
    # Exact match
    if file_key in FILE_MAP:
        return FILE_MAP[file_key]
    # Partial match
    for k, fn in FILE_MAP.items():
        if k in file_key or file_key in k:
            return fn
    return None


def scan_data_folder(base: Path) -> list[Path]:
    """Quét tất cả .xlsx file trong Data/ và các subfolder."""
    files = []
    for pattern in ["*.xlsx", "**/*.xlsx"]:
        files.extend(base.glob(pattern))
    return sorted(set(files))


# ─── Registry helpers ─────────────────────────────────────────────────────────

def get_registry(sb) -> dict[str, dict]:
    res = sb.table("data_file_registry").select("*").execute()
    return {r["file_key"]: r for r in (res.data or [])}


def update_registry(sb, file_key: str, file_name: str, sha: str, row_count: int,
                    status: str = "ok", error_msg: str = None):
    sb.table("data_file_registry").upsert({
        "file_key":      file_key,
        "file_name":     file_name,
        "sha256":        sha,
        "row_count":     row_count,
        "last_imported": NOW(),
        "status":        status,
        "error_msg":     error_msg,
    }, on_conflict="file_key").execute()


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="Re-import tất cả, bỏ qua hash check")
    args = parser.parse_args()

    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
    if not url or not key:
        print("ERROR: set SUPABASE_URL + SUPABASE_SERVICE_KEY", file=sys.stderr)
        sys.exit(1)

    if not DATA_DIR.exists():
        print(f"ERROR: Data/ folder không tồn tại tại {DATA_DIR}", file=sys.stderr)
        sys.exit(1)

    sb = create_client(url, key)
    registry = get_registry(sb)
    files = scan_data_folder(DATA_DIR)

    if not files:
        print("Không tìm thấy file .xlsx nào trong Data/ folder")
        return

    print(f"Tìm thấy {len(files)} file .xlsx\n")

    processed, skipped, errors = 0, 0, 0

    for file_path in files:
        file_key  = resolve_file_key(file_path)
        importer  = get_importer(file_key)
        sha       = sha256_file(file_path)
        prev      = registry.get(file_key, {})

        print(f"[{file_key}] {file_path.name}")

        if not importer:
            print(f"  ⚠ Không có importer cho key '{file_key}' — bỏ qua")
            skipped += 1
            continue

        if not args.force and prev.get("sha256") == sha:
            print(f"  ✓ Không thay đổi (hash match) — bỏ qua")
            skipped += 1
            continue

        try:
            row_count = importer(sb, file_path)
            update_registry(sb, file_key, file_path.name, sha, row_count)
            print(f"  ✅ Import thành công — {row_count} records")
            processed += 1
        except Exception as e:
            err = str(e)[:200]
            update_registry(sb, file_key, file_path.name, sha, 0, status="error", error_msg=err)
            print(f"  ❌ Lỗi: {err}")
            errors += 1

    print(f"\nHoàn thành — imported: {processed}, skipped: {skipped}, errors: {errors}")


if __name__ == "__main__":
    main()
