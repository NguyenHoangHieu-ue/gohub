"""
Import WORLDMOVE product catalog from CSV + APN Excel into ncc_products table.

Usage:
  cd D:/Kien_Thuc/Work/gohub/LamViec/HeThong
  set SUPABASE_URL=https://wfuigmfnfcijkvylrwzz.supabase.co
  set SUPABASE_SERVICE_KEY=<key from key.txt>
  python scripts/import_worldmove.py
"""
import csv
import os
import re
import sys
import openpyxl
from supabase import create_client

VENDOR    = "WORLDMOVE"
CSV_PATH  = "VENDOR/WORLDMOVE/WORLDMOVE1.csv"
APN_PATH  = "VENDOR/WORLDMOVE/eSIM apn April.xlsx"
BATCH_SIZE = 500


# ─── Parsing helpers ──────────────────────────────────────────────────────────

def parse_product_name(name: str) -> dict:
    """Extract days, data_gb, is_daily, is_unlimited, throttle_kbps."""
    result = {"days": None, "data_gb": None, "is_daily": False,
              "is_unlimited": False, "throttle_kbps": None}

    m = re.search(r"(\d+)\s*[Dd]ays?", name)
    if m:
        result["days"] = int(m.group(1))

    if re.search(r"[Tt]itanium\s+AYCE", name):
        result["is_unlimited"] = True
        return result

    if re.search(r"[Pp]remium\s+[Uu]nlimited", name):
        result.update(is_unlimited=True, data_gb=1.0, is_daily=True, throttle_kbps=10_000)
        return result

    if re.search(r"[Uu]nlimited", name):
        result.update(is_unlimited=True, data_gb=2.0, is_daily=True, throttle_kbps=5_000)
        return result

    m = re.search(r"([\d.]+)\s*GB\s*(/day)?", name, re.IGNORECASE)
    if m:
        result["data_gb"] = float(m.group(1))
        result["is_daily"] = m.group(2) is not None

    if result["data_gb"] is None:
        m = re.search(r"([\d.]+)\s*MB\s*(/day)?", name, re.IGNORECASE)
        if m:
            result["data_gb"] = round(float(m.group(1)) / 1000, 4)
            result["is_daily"] = m.group(2) is not None

    m = re.search(r"(\d+)\s*kbps", name, re.IGNORECASE)
    if m:
        result["throttle_kbps"] = int(m.group(1))
    else:
        m = re.search(r"(\d+(?:\.\d+)?)\s*[Mm]bps", name, re.IGNORECASE)
        if m:
            result["throttle_kbps"] = int(float(m.group(1)) * 1000)

    if result["throttle_kbps"] is None:
        result["throttle_kbps"] = 128

    return result


def _clean(v) -> str | None:
    return str(v).strip() if v else None


def parse_apn_file(xlsx_path: str) -> dict:
    """
    Parse eSIM APN Excel (row 28 = header, rows 29+ = data).

    Returns: dict mapping region_hint (str) → apn_data (dict).
    Multiple carrier rows per plan are joined into apn_telecom_providers.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["工作表1"]

    plans: dict = {}
    current_hint: str | None = None
    current_data: dict = {}
    current_carriers: list = []

    def _save():
        if current_hint is not None:
            d = dict(current_data)
            d["apn_telecom_providers"] = "\n".join(c for c in current_carriers if c) or None
            plans[current_hint] = d

    for r in range(29, ws.max_row + 1):
        cols = [ws.cell(r, c).value for c in range(1, 11)]
        plan_name, prepaid, _local, telecom, network, apn, roaming, coverage, notif, data_reset = cols

        if plan_name:
            _save()

            # Extract region hint: take the part after the first line ("Worldmove..." prefix)
            lines = str(plan_name).split("\n")
            # First line contains "Worldmove" + optional variant; second line (if any) is the region
            if len(lines) >= 2:
                hint = lines[-1].strip()
                # Strip trailing single-char variant labels: " A", " B"
                hint = re.sub(r"\s+[A-Z]$", "", hint).strip()
            else:
                # No newline — try to extract after "Worldmove " prefix
                hint = re.sub(r"^Worldmove\s*[A-Za-z]*\s*[,\n]?\s*", "", lines[0]).strip()
                if not hint:
                    hint = lines[0].strip()

            current_hint = hint
            current_data = {
                "apn":               _clean(apn),
                "apn_network_type":  _clean(network),
                "apn_roaming_carrier": _clean(roaming),
                "apn_coverage_area": _clean(coverage),
                "apn_notification":  _clean(notif),
                "apn_data_reset":    _clean(data_reset),
                "apn_prepaid_card":  _clean(prepaid),
            }
            current_carriers = [telecom.strip()] if telecom and telecom.strip() else []
        elif telecom and telecom.strip():
            current_carriers.append(telecom.strip())

    _save()
    print(f"  Parsed {len(plans)} APN plan entries from {xlsx_path}")
    return plans


def find_apn(region: str | None, apn_lookup: dict) -> dict:
    """
    Best-effort match: CSV region → APN plan data.
    Strategy:
      1. Exact match (case-insensitive)
      2. APN hint is a substring of region (or vice versa)
      3. Pick the longest-matching hint (most specific)
    """
    if not region:
        return {}

    region_l = region.strip().lower()

    # Exact
    for hint, data in apn_lookup.items():
        if hint.lower() == region_l:
            return data

    # Substring — collect all candidates, pick longest hint
    candidates = []
    for hint, data in apn_lookup.items():
        hint_l = hint.lower()
        if hint_l in region_l or region_l in hint_l:
            candidates.append((len(hint), data))

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    return {}


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
    if not url or not key:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set", file=sys.stderr)
        sys.exit(1)

    sb = create_client(url, key)

    # Load APN data
    print("Loading APN data...")
    apn_lookup = parse_apn_file(APN_PATH)

    # Load CSV
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    print(f"Loaded {len(rows)} rows from {CSV_PATH}")

    batch = []
    skipped = 0
    apn_matched = 0

    for r in rows:
        wm_id = (r.get("wmproductId") or "").strip()
        if not wm_id:
            skipped += 1
            continue

        name   = (r.get("product name") or "").strip()
        region = (r.get("region") or "").strip() or None
        parsed = parse_product_name(name)

        cogs_raw = (r.get("cost price (NT)") or "").strip()
        cogs = float(cogs_raw) if cogs_raw else None

        apn_data = find_apn(region, apn_lookup)
        if apn_data:
            apn_matched += 1

        batch.append({
            "vendor":              VENDOR,
            "vendor_product_id":   wm_id,
            "vendor_internal_id":  (r.get("productId") or "").strip() or None,
            "product_name":        name or None,
            "region":              region,
            "sim_type":            (r.get("type") or "").strip() or None,
            "days":                parsed["days"],
            "data_gb":             parsed["data_gb"],
            "is_daily":            parsed["is_daily"],
            "is_unlimited":        parsed["is_unlimited"],
            "throttle_kbps":       parsed["throttle_kbps"],
            "cogs":                cogs,
            "cogs_currency":       "TWD",
            "is_kyc":              False,
            "is_lesim":            (r.get("leSIM") or "").strip().upper() == "Y",
            "status":              "active",
            # APN data (may be None if no match)
            **apn_data,
        })

        if len(batch) >= BATCH_SIZE:
            sb.table("ncc_products").upsert(
                batch, on_conflict="vendor,vendor_product_id"
            ).execute()
            print(f"  Upserted {len(batch)} rows...")
            batch = []

    if batch:
        sb.table("ncc_products").upsert(
            batch, on_conflict="vendor,vendor_product_id"
        ).execute()
        print(f"  Upserted {len(batch)} rows (final batch)")

    print(f"Done. Skipped {skipped} rows with empty wmproductId.")
    print(f"APN matched: {apn_matched}/{len(rows) - skipped} products ({apn_matched*100//(len(rows)-skipped)}%)")


if __name__ == "__main__":
    main()
