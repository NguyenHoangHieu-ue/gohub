"""
Import country reference data into Supabase.

Usage:
  cd D:/Kien_Thuc/Work/gohub/LamViec/HeThong
  set SUPABASE_URL=https://wfuigmfnfcijkvylrwzz.supabase.co
  set SUPABASE_SERVICE_KEY=<key>
  python scripts/import_countries.py
"""
import os, sys
import openpyxl
from supabase import create_client

COUNTRIES_FILE        = "TaiLieuCongTy_Chung/countries-2026-06-05.xlsx"
SUPPORT_COUNTRIES_FILE = "TaiLieuCongTy_Chung/support-countries-2026-06-05.xlsx"


def main():
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
    if not url or not key:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set", file=sys.stderr)
        sys.exit(1)
    sb = create_client(url, key)

    # ── ref_countries ────────────────────────────────────────────────────────
    wb1 = openpyxl.load_workbook(COUNTRIES_FILE, data_only=True)
    ws1 = wb1.active
    countries = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        code, name, name_vn = row[0], row[1], row[2] if len(row) > 2 else None
        if not code:
            continue
        countries.append({
            "code":    str(code).strip(),
            "name":    str(name).strip() if name else "",
            "name_vn": str(name_vn).strip() if name_vn else None,
        })

    sb.table("ref_countries").delete().neq("code", "").execute()
    sb.table("ref_countries").insert(countries).execute()
    print(f"Imported {len(countries)} countries.")

    # ── ref_support_countries ────────────────────────────────────────────────
    wb2 = openpyxl.load_workbook(SUPPORT_COUNTRIES_FILE, data_only=True)
    ws2 = wb2.active
    support = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        code = row[0]; name = row[1]
        sc = row[2] if len(row) > 2 else None
        sc_vn = row[3] if len(row) > 3 else None
        cc = row[4] if len(row) > 4 else None
        if not code:
            continue
        support.append({
            "code":               str(code).strip(),
            "name":               str(name).strip() if name else str(code).strip(),
            "support_country":    str(sc).strip() if sc else None,
            "support_country_vn": str(sc_vn).strip() if sc_vn else None,
            "country_codes":      str(cc).strip() if cc else None,
        })

    sb.table("ref_support_countries").delete().neq("code", "").execute()
    sb.table("ref_support_countries").insert(support).execute()
    print(f"Imported {len(support)} support-country groups.")


if __name__ == "__main__":
    main()
